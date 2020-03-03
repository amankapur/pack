import math
import datetime
import time


#from mailmerge import MailMerge
from copy import copy

from openpyxl.styles import Alignment, Font, Border, Side, NamedStyle

from pprint import pprint

import os
import qrcode

def timeit(method):
	def timed(*args, **kw):
		ts = time.time()
		result = method(*args, **kw)
		te = time.time()
		if 'log_time' in kw:
			name = kw.get('log_name', method.__name__.upper())
			kw['log_time'][name] = int((te - ts) * 1000)
		else:
			print '%r  %2.2f ms' % (method.__name__, (te - ts) * 1000)
		return result
	return timed

## takes 1000000 and returns 10,00,000
def commafiy(input):
  input_list = list(str(input))
  if len(input_list) <= 1:
      formatted_input = input
  else:
      first_number = input_list.pop(0)
      last_number = input_list.pop()
      formatted_input = first_number + (
          (''.join(l + ',' * (n % 2 == 1) for n, l in enumerate(reversed(input_list)))[::-1] + last_number)
      )
      if len(input_list) % 2 == 0:
          formatted_input.lstrip(',')
  return formatted_input

def create_dir(out_dir):
	if not os.path.exists(out_dir):
		os.makedirs(out_dir)
	return out_dir


# s is qr code text
def create_qr_code(s, options={}):
	opts = {
		'version': 1,
		'error_correction': qrcode.constants.ERROR_CORRECT_H,
		'box_size': 10,
		'border': 1,
		'fit': True,
		'ext': 'jpg'
	}

	for k,v in options.iteritems():
		opts[k] = v


	out_dir = create_dir(os.path.abspath(os.path.dirname(__file__)) + "/qr_images/")

	dir_s = s.replace('/', '_')
	p = out_dir + '%s.%s' % (dir_s, opts['ext'])

	if os.path.isfile(p):
		return p

	qr = qrcode.QRCode(
		version = opts['version'],
		error_correction = opts['error_correction'],
		box_size = opts['box_size'],
		border = opts['border'],
	)

	qr.add_data(s)
	qr.make(fit=opts['fit'])
	img = qr.make_image()

	img.save(p)
	return p


# needs ssconvert installed
def convert_to_xls(in_path):
	os.system("ssconvert %s %s" % (in_path, in_path[:-1]))

def get_min(arr_of_dicts, key):
	min_val = float("inf")

	for d in arr_of_dicts:
		min_val = min(min_val, d[key])
	return min_val

def get_max(arr_of_dicts, key):
	max_val = float("-inf")

	for d in arr_of_dicts:
		max_val = max(max_val, d[key])
	return max_val


#keymap is a has str->str
# it maps keys of the in_doc merge fields to data keys/funcs
# data is [{..},{...}] format
#def my_merge(in_doc_path, out_doc_path, data, key_map=None):
#    mm_data = []
#    if key_map:
#        for d in data:
#            mm_d = {}
#
#            for mm_key, d_key in key_map.iteritems():
#                v = ''
#                if callable(d_key):
#                    v = d_key(d)
#                else:
#                    v = d[d_key]
#
#                mm_d[mm_key] = v
#
#            mm_data.append(mm_d)
#    else:
#        mm_data = data
#
#    formatted_d = [{},{}]
#    for d in mm_data:
#        for k,v in d.iteritems():
#            d[k] = str(v)
#        formatted_d.append(d)
#
#    MAX_SIZE_PER_FILE = 5000
#    with MailMerge(in_doc_path) as document:
#        document.merge_pages(formatted_d[:MAX_SIZE_PER_FILE])
#        document.write(out_doc_path)


CELL_STYLES = {
	'HEADER':{
		'alignment': Alignment(horizontal='center', vertical='center', wrap_text="true"),
		'font': Font(bold=True),
		'border':  Border(left=Side(style='thin'),
											 right=Side(style='thin'),
											 top=Side(style='thin'),
											 bottom=Side(style='thin'))
	},
	'TEXT':{
		'alignment': Alignment(horizontal='center', vertical='center'),
		'border':  Border(left=Side(style='thin'),
											 right=Side(style='thin'),
											 top=Side(style='thin'),
											 bottom=Side(style='thin'))
	},
	'TITLE':{
		'alignment': Alignment(horizontal='center', vertical='center'),
		'font': Font(bold=True, size=15)
	},
	'DATE_STYLE': NamedStyle(name='datetime', number_format='DD/MM/YYYY')
}

def insert_row(sheet, data, row_num, styles={}):
	col_idx = 1
	for d in data:
		cell = sheet.cell(row=row_num, column=col_idx, value=d or "")

		if isinstance(d, datetime.datetime) or isinstance(d, datetime.date):
			cell.style = CELL_STYLES['DATE_STYLE']

		if "alignment" in styles.keys():
			cell.alignment = styles["alignment"]
		if "font" in styles.keys():
			cell.font = styles["font"]
		if 'border' in styles.keys():
			cell.border = styles['border']

		col_idx +=1


def copy_style(new_cell, cell):
	new_cell.font = copy(cell.font)
	new_cell.border = copy(cell.border)
	new_cell.fill = copy(cell.fill)
	new_cell.number_format = copy(cell.number_format)
	new_cell.protection = copy(cell.protection)
	new_cell.alignment = copy(cell.alignment)


def get_str(k):
	if type(k) == datetime.datetime:
		d = str(k.day)
		if len(d) < 2:
			d = '0' + d

		m = str(k.month)
		if len(m) < 2:
			m = '0' + m

		y = str(k.year)
		if len(y) > 2:
			y = y[2:]

		return '%s/%s/%s'%(d,m,y)
	else:
		return str(k)

def display_from_to(from_val, to_val, splitter="-"):
	if from_val is None or to_val is None:
		return None
	if from_val == to_val:
		return str(from_val)
	else:
		return str(from_val) + ' %s ' % splitter + str(to_val)

def display_time(t):
	a = ' AM'
	if t.hour > 12:
		a = ' PM'
	return str(t.hour%12) + ':' + str(t.minute) + a

def display_month(m):
	if m == 1:
		return 'January'
	if m == 3:
		return 'March'
	if m == 2:
		return 'February'
	if m == 4:
		return 'April'
	return m

def round_up(x, n):
	n = float(n)
	return int(math.ceil(x / n) * n)

def get_string_sub_code(code):
	c = str(code)
	if len(c) != 2:
		c = '0'+ c
	return c

def zerofy(n):
	if n < 10:
		return '0' + str(n)
	else:
		return str(n)

def time_plus(time, timedelta):
	start = datetime.datetime(
			2000, 1, 1,
			hour=time.hour, minute=time.minute, second=time.second)
	end = start + timedelta
	return end.time()



def split_arr_n(arr, x):
	n = int(math.ceil(len(arr)/float(x)))

	new_arr = []
	for i in range(0, n):
		for j in range(0, x):
			new_index = i+j*n
			if new_index < len(arr):
				new_arr.append(arr[new_index])
			# else:
			#   new_arr.append(None)

	return new_arr

def sort_by(items, columns):
	return multikeysort(items, columns)

def multikeysort(items, columns):
	from operator import itemgetter
	comparers = [ ((itemgetter(col[1:].strip()), -1) if col.startswith('-') else (itemgetter(col.strip()), 1)) for col in columns]
	def comparer(left, right):
			for fn, mult in comparers:
					result = cmp(fn(left), fn(right))
					if result:
							return mult * result
			else:
					return 0
	return sorted(items, cmp=comparer)

def group_by(input_data, keys, expect_single=False):
	def _group(datum, k):
		data = {}
		for d in datum:
			v = None
			if callable(k):
				v = k(d)
			else:
				v = d[k]
			# if v:
			# if v not in data.keys():
			if v not in data:
				if expect_single:
					data[v] = d
				else:
					data[v] = [d]
			elif expect_single:
				pprint(d)
				pprint(data[v])
				raise ValueError('Expected single item in group_by')
			else:
				data[v].append(d)
		return data

	if type(keys) in [str, unicode] or callable(keys):
		return _group(input_data, keys)

	data = _group(input_data, keys[0])
	for k, datum in data.iteritems():
		data[k] = _group(datum, keys[1])

	return data


def replace_all(string, d):
	s = string
	for k in reversed(sorted(d.keys(), key=len)):
		v = d[k]
		s = s.replace(k, v)
	return s

def str_grouping(arr, splitter="-"):
	if arr == None or len(arr) == 0:
		return []
	arr = list(set(arr))
	arr.sort()

	groups = []

	i = 0
	g = []
	group_start = arr[i]
	while i < len(arr):
		if i + 1 < len(arr):
			if arr[i+1] - arr[i] != 1:
				g.append(display_from_to(group_start, arr[i], splitter))
				group_start = arr[i+1]
		else:
			g.append(display_from_to(group_start, arr[i], splitter))

		i += 1


	return g

def str_str_grouping(data, key=None, joiner=",", splitter="-"):
	s = joiner	+ " "
	arr = [d[key] if key else d for d in data]
	return s.join(str_grouping(arr, splitter))


class Cache():
	def __init__(self):
		self._data = {}

	def add(self, key, value):
		self._data[key] = value
		return value

	def get(self, key):
		if key in self._data:
			return self._data[key]
		return None

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill
