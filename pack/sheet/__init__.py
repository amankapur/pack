import datetime
from operator import itemgetter
from pprint import pprint
from copy import copy

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, NamedStyle
from openpyxl.worksheet.pagebreak import Break, PageBreak
from openpyxl.utils import get_column_letter

from pack.utils import get_str, str_grouping, style_range


CELL_STYLES = {
	'HEADER':{
		'alignment': Alignment(horizontal='center', vertical='center', wrap_text="true"),
		'font': Font(bold=True),
		'border':  Border(left=Side(style='thin'),
											 right=Side(style='thin'),
											 top=Side(style='thin'),
											 bottom=Side(style='thin'))
	},
	'HEADER_TOP':{
		'alignment': Alignment(horizontal='center', vertical='center', wrap_text="true"),
		'font': Font(bold=True, size=15),
		'border':  Border(left=Side(style='thin'),
											 right=Side(style='thin'),
											 top=Side(style='thin'),
											 bottom=Side(style='thin'))
	},
	'TEXT':{
		'alignment': Alignment(horizontal='center', vertical='center'),
		'border':  Border(left=Side(style='thin'),
											 right=Side(style='thin'),
											 top=Side(style='thin'),
											 bottom=Side(style='thin'))
	},
	'TITLE':{
		'alignment': Alignment(horizontal='center', vertical='center'),
		'font': Font(bold=True, size=15)
	},
	'DATE_STYLE': NamedStyle(name='datetime', number_format='DD/MM/YYYY')
}


class Sheet():

	def __init__(self, out_file, options={}):
		self.data = []
		self.headers = []
		self.sort_order = []
		self.out_file = out_file

		self.options = {
			'xls': False,
			'title': None,
			'autosize': True,
			'freeze_headers': True,
			'sub_total_page_break': True
		}

		for k,v in options.iteritems():
			self.options[k] = v

		self._wb = None
		self._ws = None
		if 'workbook' in self.options:
			self._wb = self.options['workbook']
			self._wb.create_sheet(self.options['title'])
			self._ws = self._wb[self.options['title']]
		else:
			self._wb = Workbook()
			self._ws = self._wb.active
			if self.options['title']:
				self._ws.title = options['title']


		self._max_width = {}
		self._sub_total_idx = None
		self._sub_total_funcs = {}
		self._sub_total_data = {}

		self._grand_total_props = None
		self._group_cols = {}
		self.breaks = []

	def _set_cell(self, value, row_idx, col_idx, style="TEXT"):
		cell = self._ws.cell(row=row_idx, column=col_idx+1, value=value or "")

		if isinstance(value, datetime.datetime) or isinstance(value, datetime.date):
			cell.style = CELL_STYLES['DATE_STYLE']

		styles = CELL_STYLES[style]

		if "alignment" in styles.keys():
			cell.alignment = styles["alignment"]
		if "font" in styles.keys():
			cell.font = styles["font"]
		if 'border' in styles.keys():
			cell.border = styles['border']

		width = len(str(value)) if type(value) != unicode else 10
		if style != "HEADER_TOP":
			if col_idx not in self._max_width.keys():
				self._max_width[col_idx] = width
			else:
				self._max_width[col_idx] = max(self._max_width[col_idx], width)

		return cell

	def _insert_row(self, row_data, row_idx, style):
		if style == 'HEADER_TOP':
			self._ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(self.headers[-1]))
			self._set_cell(row_data[0], row_idx, 0, style)

			rang = 'A%s:%s%s' % (str(row_idx), get_column_letter(len(self.headers[-1])), str(row_idx))

			s = CELL_STYLES[style]
			style_range(self._ws, rang, border=s['border'], font=s['font'], alignment=s['alignment'])

		else:
			for col_idx, cell_val in enumerate(row_data):
				if type(cell_val) == list:
					cell_val = ", ".join([str(s) for s in cell_val])
				self._set_cell(cell_val, row_idx, col_idx, style)

	def _update_sub_total_data(self, row_data):
		if self._sub_total_idx != None:
			for sub_idx, sub_func in self._sub_total_funcs.iteritems():
				if sub_idx not in self._sub_total_data.keys():
					self._sub_total_data[sub_idx] = 0
					if sub_func == 'count_uniq':
						self._sub_total_data[sub_idx] = []

				val = row_data[sub_idx]
				if val == None:
					val = 0

				if type(val) in [str, unicode] and len(val.strip())==0:
					val = 0

				if sub_func == 'count':
					self._sub_total_data[sub_idx] += 1
				elif sub_func == 'sum':
					self._sub_total_data[sub_idx] += val
				elif sub_func == 'min':
					self._sub_total_data[sub_idx] = min(self._sub_total_data[sub_idx], val)
				elif sub_func == 'max':
					self._sub_total_data[sub_idx] = max(self._sub_total_data[sub_idx], val)
				elif sub_func == 'count_uniq':
					if type(val) in [str, unicode]:
						if '-' in val:
							val = range(int(val.split("-")[0]), int(val.split("-")[1])+1)
						else:
							val = [int(val)]
					else:
						val = [val]
					self._sub_total_data[sub_idx] = list(set(self._sub_total_data[sub_idx]+val))

				self._sub_total_data[self._sub_total_idx] = row_data[self._sub_total_idx]

	def _insert_subtotal_row(self, row_idx, page_break):
		r = []
		for i in range(len(self.headers[-1])):
			r.append("")

		for idx, val in self._sub_total_data.iteritems():
			if idx == self._sub_total_idx:
				r[idx] = get_str(val) + ' Total'
			else:
				if type(val) == list:
					# r[idx] = len(val)
					# pprint(val)
					r[idx] = str(len(val))+ ' Total' + "\n (" + ", ".join(str_grouping([int(v) for v in val])) + ")"
				else:
					r[idx] = val

		self._insert_row(r, row_idx, 'HEADER')
		if page_break:
			# self._ws.page_breaks.append(Break(id=row_idx))
			self.breaks.append(Break(id=row_idx))
		self._sub_total_data = {}

	def _insert_grandtotal_row(self, row_idx):
		r = []
		for i in self.headers[-1]:
			r.append("")

		r[1] = 'Grand Total'

		sorted_data = self._get_sorted_data()
		for col_idx, grand_total_func in self._grand_total_props.iteritems():
			val_array = [d[col_idx] for d in sorted_data]
			val_array = [0 if k == None else k for k in val_array]
			if grand_total_func == 'sum':
				val = sum(val_array)
			elif grand_total_func == 'min':
				val = min(val_array)
			elif grand_total_func == 'max':
				val = max(val_array)
			elif grand_total_func == 'count':
				val = len(val_array)
			elif grand_total_func == 'count_uniq':
				big_arr = []
				for v in val_array:
					if type(v) in [str, unicode] and '-' in v:
						big_arr += range(int(v.split("-")[0]), int(v.split("-")[1])+1)
					else:
						big_arr.append(v)

				val = str(len(big_arr))+ ' Total' + "\n (" + ", ".join(str_grouping([int(v) for v in big_arr])) + ")"

			r[col_idx] = val

		self._insert_row(r, row_idx, 'TITLE')


	def _get_sorted_data(self):
		sorted_data = self.data

		if len(self.sort_order) > 0:
			for i in reversed(self.sort_order):
				v = i
				reverse = False
				if i < 0:
					v = -1 * i
					reverse = True
				sorted_data.sort(key=itemgetter(v), reverse=reverse)

		return sorted_data

	def _insert_headers(self):
		r_idx = 1
		for i, h in enumerate(self.headers):
			if i != len(self.headers) - 1 and len(self.headers)>1:
				self._insert_row(h, r_idx, 'HEADER_TOP')
			else:
				self._insert_row(h, r_idx, 'HEADER')
			r_idx += 1
		return r_idx

	def _insert_data(self, r_idx):
		sub_total_val = None
		sorted_data = self._get_sorted_data()
		has_subtotal = self._sub_total_idx != None
		for i, row_data in enumerate(sorted_data):
			if has_subtotal:
				if sub_total_val == None:
					sub_total_val = row_data[self._sub_total_idx]
				if sub_total_val != row_data[self._sub_total_idx]:
					self._insert_subtotal_row(r_idx, self.options['sub_total_page_break'])
					r_idx += 1
					sub_total_val = row_data[self._sub_total_idx]

			if i != 0:
				prev_row_data = sorted_data[i-1]
				for group_col_idx, group_col_val in self._group_cols.iteritems():
					if row_data[group_col_idx] == group_col_val:
						row_data[group_col_idx] = None
					else:
						self._group_cols[group_col_idx] = row_data[group_col_idx]
			else:
				for group_col_idx, group_col_val in self._group_cols.iteritems():
					self._group_cols[group_col_idx] = row_data[group_col_idx]

			self._insert_row(row_data, r_idx, 'TEXT')

			if has_subtotal:
				self._update_sub_total_data(row_data)
				if i == len(sorted_data)-1:
					r_idx += 1
					self._insert_subtotal_row(r_idx, False)
					r_idx += 1

			r_idx += 1

		if self._grand_total_props != None:
			self._insert_grandtotal_row(r_idx)
			r_idx += 1
		return r_idx

	def _set_print_headers(self):
		if len(self.headers) > 0:
			self._ws.print_title_rows = '1:%s' % str(len(self.headers))

	def _auto_size_columns(self):
		if self.options['autosize']:
			for col_idx, max_w in self._max_width.iteritems():
				letter = get_column_letter(col_idx+1)
				self._ws.column_dimensions[letter].width = max_w

	def _freeze_headers(self):
		if self.options['freeze_headers']:
			self._ws.freeze_panes = "A%s" % str(len(self.headers)+1)


	def _create(self):
		print 'Creating sheet  %s ' % self.out_file
		self._wb.save(self.out_file)


	# needs ssconvert installed
	def _xls(self):
		if self.options['xls']:
			p = self.out_file
			os.system("ssconvert %s %s" % (p, p[:-1]))


	def add_headers(self, headers):
		self.headers = headers[0:-1]

		last_header = headers[-1]
		last_header_text = []

		s_hsh = {}

		for col_idx,h in enumerate(last_header):
			if type(h) == tuple:
				text, d = h
				last_header_text.append(text)

				sort_idx = d
				if type(d) != int:
					if 'sort' in d:
						sort_idx = d['sort']
					else:
						sort_idx=None

				if sort_idx:
					s_hsh[sort_idx] = col_idx

				if type(d) == dict:
					if 'sub_total' in d:
						self._sub_total_idx = col_idx

					if 'sub_total_func' in d:
						self._sub_total_funcs[col_idx] = d['sub_total_func']

					if 'grand_total_func' in d:
						if self._grand_total_props == None:
							self._grand_total_props = {}
						self._grand_total_props[col_idx] = d['grand_total_func']

					if 'group_col' in d:
						self._group_cols[col_idx] = None

			else:
				last_header_text.append(h)

		for i in range(len(last_header)+1):
			if i in s_hsh.keys():
				self.sort_order.append(s_hsh[i])
			elif -1*i in s_hsh.keys():
				self.sort_order.append(-1*s_hsh[-1*i])

		self.headers.append(last_header_text)


	def add_rows(self, rows):
		self.data = rows

	def save(self, output=True):
		r_idx = self._insert_headers()
		self._insert_data(r_idx)

		self._set_print_headers()
		self._auto_size_columns()
		self._freeze_headers()

		if output:
			self._create()
			self._xls()

		return self
