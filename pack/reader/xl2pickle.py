import os
import pickle
import logging

from pprint import pprint
from openpyxl import load_workbook


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class xl2pickle(object):

	def __init__(self, in_path, out_path, columns, use_all_sheets=False, header_rows=1):

		if not os.path.isdir(in_path) and not os.path.isfile(in_path):
			raise ValueError('Input path is not a valid directory')

		self.in_path = in_path
		self.out_path = out_path
		self.columns = columns
		self.use_all_sheets = use_all_sheets
		self.header_rows = header_rows

	def _get_formatter(self, t):
		if t == int:
			def f(value):
				return int(float(str(value)))
			return f
		if t == float:
			def f(value):
				return float(str(value))
			return f

	def _get_merged_hash(self, ws):
		ranges = {}
		for c_r in ws.merged_cells.ranges:
			if c_r.min_col == c_r.max_col:
				ranges[(c_r.min_col, c_r.min_row)] = {
					'min_row': c_r.min_row,
					'max_row': c_r.max_row,
					'jump': c_r.max_row - c_r.min_row + 1
				}
		return ranges

	def _read_file(self, f):
		logger.info('Reading from %s' % (f))
		data = []
		wb = load_workbook(f, data_only=True)
		for sh_name in wb.sheetnames:
			ws = wb[sh_name]
			r_i = self.header_rows+1
			m_hash = self._get_merged_hash(ws)
			while r_i < ws.max_row+1:
				row_data = {}
				row_jump = 1
				for col in self.columns:

					if 'merged_rows' in col:
						k = (col['index'], r_i)
						if k not in m_hash:
							pprint(m_hash)
						row_jump = m_hash[k]['jump']

					vals = None
					for i in range(row_jump):
						cell = ws.cell(row=r_i+i, column=col['index'])
						if cell.value != None:
							try:
								v = str(cell.value)
							except:
								logger.warning('Can\'t convert %s to string at %s' % (cell.value, cell.coordinate))

							if vals == None:
								vals = [cell.value]
							else:
								vals.append(cell.value)

					if vals:
						if 'formatter' in col.keys():
							vals = [col['formatter'](val) for val in  vals]
						elif 'type' in col.keys():
							vals = [self._get_formatter(col['type'])(val) for val in vals]
						else:
							k = []
							for val in vals:
								if type(val) in [str, unicode]:
									k.append(val.strip())
								else:
									k.append(val)
							vals = k

					if vals and len(vals) == 1:
						vals = vals[0]
					row_data[col['name']] = vals

				data.append(row_data)
				r_i += row_jump

			if not self.use_all_sheets:
				break
		return data


	def read_xlsx(self, save=True):
		if os.path.isdir(self.in_path):
			for fn in os.listdir(self.in_path):
				f = self.in_path + fn
				if os.path.isfile(f) and f.endswith('.xlsx') and not fn.startswith('~'):
					data = self._read_file(f)
					if save:
						p = self.out_path+'%s.p' % fn.split('.')[0]
						self._write_data(p, data)
		else:
			data = self._read_file(self.in_path)
			if save:
				self._write_data(self.out_path, data)

	def _write_data(self, p, data):
		logger.info('Writing to : ' + p)
		with open(p, 'wb') as f:
			pickle.dump(data, f)
